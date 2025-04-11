import os
import cv2
import requests
import json
import base64
import pandas as pd
import logging
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from concurrent.futures import ThreadPoolExecutor, as_completed
from PIL import Image as PilImage
from decimal import Decimal, ROUND_DOWN
# ====== 第一步：代码段 1 ======

# 图片所在的文件夹路径
input_folder = r'D:\baogao\hongzhan\ping'  # 修改为你自己的文件夹路径
output_folder = r'D:\baogao\hongzhan\jietu'  # 输出文件夹

# 创建输出文件夹（如果不存在）
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# 遍历文件夹中的所有图片
for filename in os.listdir(input_folder):
    # 检查文件是否是图片
    if filename.endswith(('.jpg', '.png', '.jpeg')):
        image_path = os.path.join(input_folder, filename)

        # 读取图片
        image = cv2.imread(image_path)

        # 检查图像是否加载成功
        if image is None:
            print(f"无法加载图片: {filename}")
            continue

        # 获取图像的尺寸
        height, width, _ = image.shape
        print(f"处理图像: {filename}, 尺寸: {width}x{height}")

        # 假设区域坐标
        region_1 = (575, 98, 623, 117)  # 左上角(371,25)，右下角(436,50)
        region_2 = (577, 122, 613, 137)  # 左上角(371,50)，右下角(436,75)

        # 提取每个区域
        region_rsrp = image[region_1[1]:region_1[3], region_1[0]:region_1[2]]
        region_sinr = image[region_2[1]:region_2[3], region_2[0]:region_2[2]]

        # 检查提取的区域是否为空（如果有必要）
        if region_rsrp.size == 0 or region_sinr.size == 0:
            print(f"区域提取失败: {filename}, 至少有一个区域为空")
        else:
            # 保存提取的区域，使用原始文件名作为前缀
            base_name = os.path.splitext(filename)[0]  # 去除扩展名
            cv2.imwrite(os.path.join(output_folder, f'{base_name}_delay.jpg'), region_rsrp)
            cv2.imwrite(os.path.join(output_folder, f'{base_name}_jitter.jpg'), region_sinr)
            print(f"处理并保存图片: {filename}")

print("所有图片已处理完成")

# ====== 第二步：代码段 2 ======

# 图片所在的文件夹路径
input_folder = r'D:\baogao\hongzhan\pictuer'  # 修改为你自己的文件夹路径
output_folder = r'D:\baogao\hongzhan\jietu'  # 输出文件夹

# 创建输出文件夹（如果不存在）
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

# 遍历文件夹中的所有图片
for filename in os.listdir(input_folder):
    # 检查文件是否是图片
    if filename.endswith(('.jpg', '.png', '.jpeg')):
        image_path = os.path.join(input_folder, filename)

        # 读取图片
        image = cv2.imread(image_path)

        # 检查图像是否加载成功
        if image is None:
            print(f"无法加载图片: {filename}")
            continue

        # 获取图像的尺寸
        height, width, _ = image.shape
        print(f"处理图像: {filename}, 尺寸: {width}x{height}")

        # 假设区域坐标
        region_1 = (371, 25, 436, 50)  # 左上角(50,50)，右下角(250,150)
        region_2 = (371, 50, 436, 75)  # 左上角(300,50)，右下角(500,150)
        region_3 = (118, 710, 215, 734)  # 左上角(50,200)，右下角(250,300)
        region_4 = (253, 710, 384, 732)  # 新区域坐标 (100, 100) 到 (300, 150)

        # 提取每个区域
        region_rsrp = image[region_1[1]:region_1[3], region_1[0]:region_1[2]]
        region_sinr = image[region_2[1]:region_2[3], region_2[0]:region_2[2]]
        region_dlul = image[region_3[1]:region_3[3], region_3[0]:region_3[2]]
        region_new = image[region_4[1]:region_4[3], region_4[0]:region_4[2]]  # 提取新区域

        # 检查提取的区域是否为空
        if region_new.size == 0:
            print(f"区域提取失败: {filename}, 区域_4为空")
        else:
            # 保存提取的区域，使用原始文件名作为前缀
            base_name = os.path.splitext(filename)[0]  # 去除扩展名
            cv2.imwrite(os.path.join(output_folder, f'{base_name}_RSRP.jpg'), region_rsrp)
            cv2.imwrite(os.path.join(output_folder, f'{base_name}_SINR.jpg'), region_sinr)
            cv2.imwrite(os.path.join(output_folder, f'{base_name}_dl.jpg'), region_dlul)
            cv2.imwrite(os.path.join(output_folder, f'{base_name}_ul.jpg'), region_new)  # 保存新区域
            print(f"处理并保存图片: {filename}")

print("所有图片已处理完成")



# ====== 第三步：代码段 3 ======

# 设置日志
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

# OCR 服务器地址
OCR_URL = "http://127.0.0.1:1224/api/ocr"

# 需要处理的图片文件夹
IMAGE_FOLDER = r"D:\baogao\hongzhan\jietu"

# 保存结果的 Excel 文件
OUTPUT_EXCEL = r"D:\baogao\hongzhan\result.xlsx"

# OCR 识别配置
OCR_OPTIONS = {
    "ocr.language": r"D:\Umi-OCR_Paddle_v2.1.4.7z\Umi-OCR_Paddle_v2.1.4\UmiOCR-data\plugins\win7_x64_PaddleOCR-json\models\ch_ppocr_mobile_v2.0_cls_infer",
    "ocr.angle": False,
    "ocr.maxSideLen": 1024,
    "tbpu.parser": "multi_para",
    "data.format": "text"
}

def image_to_base64(image_path):
    """ 读取图片并转换为 Base64 字符串 """
    try:
        with open(image_path, "rb") as img_file:
            return base64.b64encode(img_file.read()).decode("utf-8")
    except Exception as e:
        logging.error(f"读取图片失败: {image_path}, 错误: {e}")
        return None

def ocr_image(image_path):
    """ 发送 OCR 识别请求，返回识别文本 """
    base64_str = image_to_base64(image_path)
    if not base64_str:
        return "图片读取失败"

    data = {
        "base64": base64_str,
        "options": OCR_OPTIONS
    }
    headers = {"Content-Type": "application/json"}

    try:
        response = requests.post(OCR_URL, data=json.dumps(data), headers=headers, timeout=10)
        response.raise_for_status()
        res_dict = response.json()

        if res_dict.get("code") == 100:
            return res_dict.get("data", "未识别到文本")
        elif res_dict.get("code") == 200:
            return res_dict.get("data", {}).get("text", "未识别到文本")
        else:
            return f"OCR 失败: {res_dict}"

    except requests.exceptions.Timeout:
        return "OCR 服务器请求超时"
    except requests.exceptions.RequestException as e:
        return f"请求出错: {e}"
    except json.JSONDecodeError:
        return "OCR 服务器返回了无效的 JSON 数据"

def clear_sheet_data(excel_path, sheet_name="Sheet1"):
    """ 清空指定 Excel 文件中指定工作表的数据 """
    try:
        wb = load_workbook(excel_path)
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
                for cell in row:
                    cell.value = None
            wb.save(excel_path)
            logging.info(f"清空了 {sheet_name} 工作表中的数据")
        else:
            logging.warning(f"工作表 {sheet_name} 不存在")
    except Exception as e:
        logging.error(f"清空工作表 {sheet_name} 数据失败: {e}")

def format_text(text):
    """ 如果文本包含数字，则保留两位小数且不进行四舍五入 """
    try:
        # 查找文本中的所有数字
        words = text.split()
        formatted_words = []
        for word in words:
            # 尝试将每个词转换为浮动数字，如果成功则保留两位小数
            try:
                # 使用 Decimal 来避免四舍五入，保留两位小数
                formatted_word = str(Decimal(word).quantize(Decimal('0.00'), rounding=ROUND_DOWN))
                formatted_words.append(formatted_word)
            except (ValueError, ArithmeticError):
                # 如果转换失败，直接保留原始词
                formatted_words.append(word)
        return " ".join(formatted_words)
    except Exception as e:
        logging.error(f"格式化文本失败: {e}")
        return text

def process_images(folder_path, output_excel, max_workers=5):
    """ 并行处理文件夹中的图片，进行 OCR 识别并保存到 Excel """
    results = []
    image_files = [f for f in os.listdir(folder_path) if f.lower().endswith((".png", ".jpg", ".jpeg", ".bmp"))]

    if not image_files:
        logging.warning("未找到符合格式的图片")
        return

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_image = {executor.submit(ocr_image, os.path.join(folder_path, img)): img for img in image_files}

        for future in as_completed(future_to_image):
            image_name = future_to_image[future]
            try:
                text = future.result()
                formatted_text = format_text(text)  # 格式化文本，保留小数点后两位
                results.append([image_name, formatted_text])
                logging.info(f"处理完成: {image_name}")
            except Exception as e:
                logging.error(f"处理 {image_name} 失败: {e}")

    # 按文件名升序排序结果
    results.sort(key=lambda x: x[0], reverse=False)

    # 清空 Sheet1 中的所有数据
    clear_sheet_data(output_excel, sheet_name="Sheet1")

    # 使用 ExcelWriter 将结果保存到指定的 sheet 中
    with pd.ExcelWriter(output_excel, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df = pd.DataFrame(results, columns=["文件名", "识别文本"])
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        logging.info(f"识别完成，结果已保存到: {output_excel} 的 Sheet1")

# 运行批量 OCR 处理
process_images(IMAGE_FOLDER, OUTPUT_EXCEL, max_workers=5)


# ====== 第四步：代码段 4 ======

# 定义图片文件夹路径
image_folder = r"D:\baogao\hongzhan\pictuer"
output_excel = r"D:\baogao\hongzhan\GLE120809R_模板.xlsx"  # 这里是你的现有文件路径

# 自定义每张图片的插入位置和大小（通过图片路径进行匹配）
image_positions = {
    "C:/Users/17691/Desktop/baogo/hongzhan/pictuer/1hd.png": {'position': 'AC6', 'width_cm': 13.9, 'height_cm': 17.0},
    "C:/Users/17691/Desktop/baogo/hongzhan/pictuer/1hu.png": {'position': 'AT6', 'width_cm': 11.8, 'height_cm': 17.0},
    "C:/Users/17691/Desktop/baogo/hongzhan/pictuer/1cd.png": {'position': 'BK6', 'width_cm': 11.8, 'height_cm': 17.0},
    "C:/Users/17691/Desktop/baogo/hongzhan/pictuer/1cu.png": {'position': 'CB6', 'width_cm': 11.8, 'height_cm': 17.0},
    "C:/Users/17691/Desktop/baogo/hongzhan/pictuer/2hd.png": {'position': 'AC31', 'width_cm': 13.9, 'height_cm': 18.0},
    "C:/Users/17691/Desktop/baogo/hongzhan/pictuer/2hu.png": {'position': 'AT31', 'width_cm': 11.8, 'height_cm': 18.0},
    "C:/Users/17691/Desktop/baogo/hongzhan/pictuer/2cd.png": {'position': 'BK31', 'width_cm': 11.8, 'height_cm': 18.0},
    "C:/Users/17691/Desktop/baogo/hongzhan/pictuer/2cu.png": {'position': 'CB31', 'width_cm': 11.8, 'height_cm': 18.0},
    "C:/Users/17691/Desktop/baogo/hongzhan/pictuer/3hd.png": {'position': 'AC56', 'width_cm': 13.9, 'height_cm': 18.2},
    "C:/Users/17691/Desktop/baogo/hongzhan/pictuer/3hu.png": {'position': 'AT56', 'width_cm': 11.8, 'height_cm': 18.2},
    "C:/Users/17691/Desktop/baogo/hongzhan/pictuer/3cd.png": {'position': 'BK56', 'width_cm': 11.8, 'height_cm': 18.2},
    "C:/Users/17691/Desktop/baogo/hongzhan/pictuer/3cu.png": {'position': 'CB56', 'width_cm': 11.8, 'height_cm': 18.2},
}

# 打开已有的工作簿
wb = load_workbook(output_excel)

# 获取Sheet2
ws = wb['SA-网络性能验收-CQT']  # 这里获取的是Sheet2工作表

# 删除所有图片（更彻底的方式）
ws._images.clear()  # 清除所有图片

# 将厘米转换为点（pt），1 厘米 ≈ 28.35 pt
def cm_to_pt(cm):
    return cm * 28.35

# 遍历字典，根据自定义的路径、位置和大小插入图片
for img_path, details in image_positions.items():
    # 获取图片路径、插入位置、宽度和高度
    position = details['position']
    width_cm = details['width_cm']
    height_cm = details['height_cm']

    # 加载图片
    img = PilImage.open(img_path)

    # 将厘米转换为点
    width_pt = cm_to_pt(width_cm)
    height_pt = cm_to_pt(height_cm)

    # 将图片插入到 Excel
    img_openpyxl = Image(img_path)
    img_openpyxl.width = width_pt  # 设置宽度
    img_openpyxl.height = height_pt  # 设置高度

    # 插入图片到指定位置
    ws.add_image(img_openpyxl, position)

# 保存工作簿
wb.save(output_excel)
print(f"图片已成功插入到 {output_excel} 的 Sheet2 中")
# ====== 第五步：代码段 5 ======

# 定义图片文件夹路径
image_folder = r"D:\baogao\hongzhan\ping"
output_excel = r"D:\baogao\hongzhan\GLE120809R_模板.xlsx"

# 自定义插入的 6 张图片及其位置和大小
image_positions = {
    r"C:\Users\17691\Desktop\baogo\hongzhan\ping\1ping32.png": {'position': 'CS6', 'width_cm': 11.7, 'height_cm': 17.3},
    r"C:\Users\17691\Desktop\baogo\hongzhan\ping\1ping2000.png": {'position': 'DJ6', 'width_cm': 11.8, 'height_cm': 17.3},
    r"C:\Users\17691\Desktop\baogo\hongzhan\ping\2ping32.png": {'position': 'CS31', 'width_cm': 11.7, 'height_cm': 18.0},
    r"C:\Users\17691\Desktop\baogo\hongzhan\ping\2ping2000.png": {'position': 'DJ31', 'width_cm': 11.8, 'height_cm': 18.0},
    r"C:\Users\17691\Desktop\baogo\hongzhan\ping\3ping32.png": {'position': 'CS56', 'width_cm': 11.7, 'height_cm': 18.2},
    r"C:\Users\17691\Desktop\baogo\hongzhan\ping\3ping2000.png": {'position': 'DJ56', 'width_cm': 11.8, 'height_cm': 18.2},
}

# 打开工作簿和指定的工作表
wb = load_workbook(output_excel)
ws = wb['SA-网络性能验收-CQT']

# 将厘米转换为点（pt），1 厘米 ≈ 28.35 pt
def cm_to_pt(cm):
    return cm * 28.35

# 插入图片
for img_name, details in image_positions.items():
    img_path = os.path.join(image_folder, img_name)
    if not os.path.exists(img_path):
        print(f"警告: 图片 {img_path} 不存在，跳过。")
        continue

    # 加载图片
    img = PilImage.open(img_path)
    width_pt = cm_to_pt(details['width_cm'])
    height_pt = cm_to_pt(details['height_cm'])

    # 插入到 Excel
    img_openpyxl = Image(img_path)
    img_openpyxl.width = width_pt
    img_openpyxl.height = height_pt
    ws.add_image(img_openpyxl, details['position'])

# 保存工作簿
wb.save(output_excel)
print(f"6 张图片已成功插入到 {output_excel} 的 SA-网络性能验收-CQT 工作表中")
# ====== 第六步：代码段 6 ======

# 定义图片文件夹路径
image_folder = r"D:\baogao\hongzhan\pictuer"
output_excel = r"D:\baogao\hongzhan\GLE120809R_模板.xlsx"

# 自定义插入的 5 张图片及其位置和大小 
image_positions = {
    r"C:\Users\17691\Desktop\baogo\hongzhan\dt\RSRP.png": {'position': 'K7', 'width_cm': 22.6, 'height_cm':10.4},
    r"C:\Users\17691\Desktop\baogo\hongzhan\dt\SINR.png": {'position': 'K26', 'width_cm': 22.6, 'height_cm': 10.4},
    r"C:\Users\17691\Desktop\baogo\hongzhan\dt\PCI.png": {'position': 'B45', 'width_cm': 27.7, 'height_cm': 10.4},
    r"C:\Users\17691\Desktop\baogo\hongzhan\dt\DL.png": {'position': 'B64', 'width_cm': 13.8, 'height_cm': 10.5},
    r"C:\Users\17691\Desktop\baogo\hongzhan\dt\UL.png": {'position': 'H64', 'width_cm': 13.8, 'height_cm': 10.5},
}

# 打开工作簿和指定的工作表
wb = load_workbook(output_excel)
ws = wb['SA-网络性能验收-DT']  # 修改工作表为 'SA-网络性能验收-DT'

# 删除工作表中的所有图片
for image in ws._images[:]:
    ws._images.remove(image)

# 将厘米转换为点（pt），1 厘米 ≈ 28.35 pt
def cm_to_pt(cm):
    return cm * 28.35

# 插入图片
for img_name, details in image_positions.items():
    img_path = os.path.join(image_folder, img_name)
    if not os.path.exists(img_path):
        print(f"警告: 图片 {img_path} 不存在，跳过。")
        continue

    # 加载图片
    img = PilImage.open(img_path)
    width_pt = cm_to_pt(details['width_cm'])
    height_pt = cm_to_pt(details['height_cm'])

    # 插入到 Excel
    img_openpyxl = Image(img_path)
    img_openpyxl.width = width_pt
    img_openpyxl.height = height_pt
    ws.add_image(img_openpyxl, details['position'])

# 保存工作簿
wb.save(output_excel)
print(f"5 张图片已成功插入到 {output_excel} 的 SA-网络性能验收-DT 工作表中")

# 删除指定文件夹中的所有文件
jietu_folder = r"D:\baogao\hongzhan\jietu"
for filename in os.listdir(jietu_folder):
    file_path = os.path.join(jietu_folder, filename)
    try:
        if os.path.isfile(file_path):
            os.remove(file_path)
            print(f"已删除文件: {file_path}")
    except Exception as e:
        print(f"删除文件 {file_path} 失败: {e}")